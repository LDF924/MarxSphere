#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""tsv2xlsx.py — 把 labeling.tsv 转成带下拉校验的 Excel 标注表(人工标注用)"""
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
src = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "data" / "ml-router" / "labeling.tsv"
out = src.with_suffix(".xlsx")

wb = Workbook()
ws = wb.active
ws.title = "标注"
ws.append(["id", "query", "rule_suggest(参考)", "human_label(填此列)", "note"])
ws["A1"].font = ws["B1"].font = ws["C1"].font = ws["D1"].font = ws["E1"].font = Font(bold=True)
for c in "ABCDE":
    ws[f"{c}1"].fill = PatternFill("solid", fgColor="DDEEFF")

with open(src, encoding="utf-8") as f:
    for line in f.read().splitlines()[1:]:
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 5:
            parts = (parts + ["", "", "", ""])[:5]
        ws.append(parts)

# 下拉校验 human_label
dv = DataValidation(type="list", formula1='"lite,standard,deep"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"D2:D{ws.max_row}")
# 列宽
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 24
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)
wb.save(out)
print("已生成:", out)
